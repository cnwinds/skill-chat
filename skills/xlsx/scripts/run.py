#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font


def emit(payload):
    print(json.dumps(payload, ensure_ascii=False), flush=True)


def safe_name(value, default):
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", value or default).strip()
    return cleaned or default


def read_request(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def load_csv_rows(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def add_chart(sheet):
    if sheet.max_row < 2 or sheet.max_column < 2:
        return

    numeric_column = None
    for column in range(2, sheet.max_column + 1):
        values = [sheet.cell(row=row, column=column).value for row in range(2, sheet.max_row + 1)]
        if values and all(isinstance(value, (int, float)) for value in values if value is not None):
            numeric_column = column
            break

    if numeric_column is None:
        return

    chart = BarChart()
    chart.title = "自动生成图表"
    chart.y_axis.title = "值"
    chart.x_axis.title = sheet.cell(row=1, column=1).value or "项目"
    data = Reference(sheet, min_col=numeric_column, min_row=1, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, f"A{sheet.max_row + 3}")


def convert_numbers(row):
    converted = []
    for value in row:
        if value is None:
            converted.append(value)
            continue
        text = str(value).strip()
        if re.fullmatch(r"-?\d+", text):
            converted.append(int(text))
        elif re.fullmatch(r"-?\d+\.\d+", text):
            converted.append(float(text))
        else:
            converted.append(text)
    return converted


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--request", required=True)
    args = parser.parse_args()

    request_payload = read_request(args.request)
    output_dir = request_payload["session"]["outputDir"]
    os.makedirs(output_dir, exist_ok=True)

    title = request_payload["input"]["arguments"].get("title") or "SkillChat 表格"
    files = request_payload["input"].get("files", [])

    emit({"type": "progress", "message": "正在创建工作簿", "percent": 15})
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = title
    summary_sheet["A1"].font = Font(size=14, bold=True)
    summary_sheet["A2"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_sheet["A4"] = "原始请求"
    summary_sheet["B4"] = request_payload["input"]["prompt"]

    data_written = False
    for file_info in files:
        if file_info["name"].lower().endswith(".csv") and os.path.exists(file_info["path"]):
            emit({"type": "progress", "message": f"正在导入 {file_info['name']}", "percent": 50})
            rows = load_csv_rows(file_info["path"])
            sheet = workbook.create_sheet(title="Data")
            for row in rows:
                sheet.append(convert_numbers(row))
            add_chart(sheet)
            data_written = True
            break

    if not data_written:
        sheet = workbook.create_sheet(title="Notes")
        sheet.append(["字段", "内容"])
        sheet.append(["标题", title])
        sheet.append(["请求", request_payload["input"]["prompt"]])
        sheet.append(["输入文件数", len(files)])

    output_name = f"{safe_name(title, 'report')}.xlsx"
    output_path = os.path.join(output_dir, output_name)

    emit({"type": "progress", "message": "正在写入 XLSX 文件", "percent": 80})
    workbook.save(output_path)

    emit({"type": "artifact", "path": os.path.relpath(output_path, request_payload["session"]["workDir"]), "label": output_name})
    emit({"type": "result", "message": "XLSX 生成完成"})


if __name__ == "__main__":
    main()
